# coding: utf-8

'''
Elaborado por Marceo Magalhães Silva de Sousa em Abril de 2019. 
Permitida a utilização deste código, desde que citada expressamente a autoria.

@autor: Marcelo Magalhães Silva de Sousa
'''


class classe_Secretaria():
    def __init__(self, nomeSecretaria):
        self.nomeSecretaria = nomeSecretaria
        contadorLinhasSecretaria = 0
        contadorLinhasSecSubordinada1 = 0  
        contadorLinhasSecSubordinada2 = 0
        contadorLinhasSecSubordinada3 = 0
        nomeSubordinada1 = ''
        nomeSubordinada2 = ''
        nomeSubordinada3 = ''
        self.contadorLinhasSecretaria = contadorLinhasSecretaria
        self.contadorLinhasSecSubordinada1 = contadorLinhasSecSubordinada1  
        self.contadorLinhasSecSubordinada2 = contadorLinhasSecSubordinada2
        self.contadorLinhasSecSubordinada3 = contadorLinhasSecSubordinada3 
        self.nomeSubordinada1 = nomeSubordinada1
        self.nomeSubordinada2 = nomeSubordinada2
        self.nomeSubordinada3 = nomeSubordinada3
 
    def __call__(self):
        return self.nomeSecretaria

    def set_Secretaria(self, nomeSecretaria):
        self.nomeSecretaria = nomeSecretaria
        
    def get_Secretaria(self):
        return self.nomeSecretaria
    
    def set_Subordinada1(self, nomeSubordinada1):
        self.nomeSubordinada1 = nomeSubordinada1
    
    def set_Subordinada2(self, nomeSubordinada2):
        self.nomeSubordinada2 = nomeSubordinada2
        
    def set_Subordinada3(self, nomeSubordinada3):
        self.nomeSubordinada3 = nomeSubordinada3
        
    def get_Subordinada1(self):
        return self.nomeSubordinada1
    
    def get_Subordinada2(self):
        return self.nomeSubordinada2
    
    def get_Subordinada3(self):
        return self.nomeSubordinada3
    
    def set_Contador_LinhasSecretaria(self, contadorLinhasSecretaria):
        self.contadorLinhasSecretaria = contadorLinhasSecretaria
        
    def get_Contador_LinhasSecretaria(self):
        return self.contadorLinhasSecretaria

    def set_Contador_LinhasSecSubordinada1(self, contadorLinhasSecSubordinada1):
        self.contadorLinhasSecSubordinada1 = contadorLinhasSecSubordinada1
        
    def get_Contador_LinhasSecSubordinada1(self):
        return self.contadorLinhasSecSubordinada1    
           
    def set_Contador_LinhasSecSubordinada2(self, contadorLinhasSecSubordinada2):
        self.contadorLinhasSecSubordinada2 = contadorLinhasSecSubordinada2
        
    def get_Contador_LinhasSecSubordinada2(self):
        return self.contadorLinhasSecSubordinada2  
    
    def set_Contador_LinhasSecSubordinada3(self, contadorLinhasSecSubordinada3):
        self.contadorLinhasSecSubordinada3 = contadorLinhasSecSubordinada3
        
    def get_Contador_LinhasSecSubordinada3(self):
        return self.contadorLinhasSecSubordinada3     


class classe_Bloco():
    def __init__(self, tituloBloco):
        self.tituloBloco = tituloBloco        
        contadorLinhasTitBloco = 0
        contadorLinhasSubTitBloco = 0  
        contadorLinhasDemaisSubTitBloco = 0
        contadorLinhasTotalBloco = 0
        tituloSubordinadoBloco1 = ''
        tituloSubordinadoBloco2 = ''
        textoInicioBloco = ''
        textoFimBloco = ''
        textoBloco = ''       
        self.contadorLinhasTitBloco = contadorLinhasTitBloco
        self.contadorLinhasSubTitBloco = contadorLinhasSubTitBloco 
        self.contadorLinhasDemaisSubTitBloco = contadorLinhasDemaisSubTitBloco 
        self.contadorLinhasTotalBloco = contadorLinhasTotalBloco
        self.tituloSubordinadoBloco1 = tituloSubordinadoBloco1
        self.tituloSubordinadoBloco2 = tituloSubordinadoBloco2    
        self.textoInicioBloco = textoInicioBloco
        self.textoFimBloco = textoFimBloco
        self.textoBloco = textoBloco
        

    def __call__(self):
        return self.tituloBloco

    def set_Titulo_Bloco(self, tituloBloco):
        self.tituloBloco = tituloBloco
        
    def get_Titulo_Bloco(self):
        return self.tituloBloco
 
    def set_Titulo_Subordinado1(self, tituloSubordinadoBloco1):
        self.tituloSubordinadoBloco1 = tituloSubordinadoBloco1
                
    def get_Titulo_Subordinado1(self):
        return self.tituloSubordinadoBloco1       

    def set_Titulo_Subordinado2(self, tituloSubordinadoBloco2):
        self.tituloSubordinadoBloco2 = tituloSubordinadoBloco2
        
    def get_Titulo_Subordinado2(self):
        return self.tituloSubordinadoBloco2          

    def set_Texto_Bloco(self, textoBloco):
        self.textoBloco = textoBloco
        
    def get_Texto_Bloco(self):
        return self.textoBloco  

    def set_Inicio_Bloco(self, textoInicioBloco):
        self.textoInicioBloco = textoInicioBloco
        
    def get_Inicio_Bloco(self):
        return self.textoInicioBloco  
    
    def set_Fim_Bloco(self, textoFimBloco):
        self.textoFimBloco = textoFimBloco
        
    def get_Fim_Bloco(self):
        return self.textoFimBloco      

    def set_Pagina_Inicio_Bloco(self, paginaInicioBloco):
        self.paginaInicioBloco = paginaInicioBloco
        
    def get_Pagina_Inicio_Bloco(self):
        return self.paginaInicioBloco   
    
    def set_Pagina_Fim_Bloco(self, paginaFimBloco):
        self.paginaFimBloco = paginaFimBloco
        
    def get_Pagina_Fim_Bloco(self):
        return self.paginaFimBloco      
    
    def set_Linha_Inicio_Bloco(self, linhaInicioBloco):
        self.linhaInicioBloco = linhaInicioBloco
        
    def get_Linha_Inicio_Bloco(self):
        return self.linhaInicioBloco
    
    def set_Linha_Fim_Bloco(self, linhaFimBloco):
        self.linhaFimBloco = linhaFimBloco
        
    def get_Linha_Fim_Bloco(self):
        return self.linhaFimBloco      
    
    def set_Contador_Linhas(self, contadorLinhasBloco):
        self.contadorLinhasBloco = contadorLinhasBloco
        
    def get_Contador_Linhas(self):
        return self.contadorLinhasBloco             
    
    def set_Contador_LinhasTitBloco(self, contadorLinhasTitBloco):
        self.contadorLinhasTitBloco = contadorLinhasTitBloco
        
    def get_Contador_LinhasTitBloco(self):
        return self.contadorLinhasTitBloco        
    
    def set_Contador_LinhasSubTitBloco(self, contadorLinhasSubTitBloco):
        self.contadorLinhasSubTitBloco = contadorLinhasSubTitBloco
        
    def get_Contador_LinhasSubTitBloco(self):
        return self.contadorLinhasSubTitBloco 
                                        
    def set_Contador_LinhasDemaisSubTitBloco(self, contadorLinhasDemaisSubTitBloco):
        self.contadorLinhasDemaisSubTitBloco = contadorLinhasDemaisSubTitBloco
        
    def get_Contador_LinhasDemaisSubTitBloco(self):
        return self.contadorLinhasDemaisSubTitBloco              

    def set_Contador_LinhasTotalBloco(self, contadorLinhasTotalBloco):
        self.contadorLinhasTotalBloco = contadorLinhasTotalBloco
        
    def get_Contador_LinhasTotalBloco(self):
        return self.contadorLinhasTotalBloco      
                                                                    
    

class classe_Diario():

    def __init__(self, numeroDiario=None):
        self.numeroDiario = numeroDiario
        contadorLinhasGeral = 0
        contadorLinhasPagina = 0    
        contadorLinhasSecao = 0    
        self.contadorLinhasGeral = contadorLinhasGeral
        self.contadorLinhasPagina = contadorLinhasPagina
        self.contadorLinhasSecao = contadorLinhasSecao
        
    def __call__(self):
        return self.numeroDiario 

    def set_Numero_Diario(self, numeroDiario):
        self.numeroDiario = numeroDiario
        
    def get_Numero_Diario(self):
        return self.numeroDiario
 
    def set_Data_Diario(self, dataDiario):
        self.dataDiario = dataDiario
                
    def get_Data_Diario(self):
        return self.dataDiario       
    
    def set_Data_Diario_Extenso(self, dataDiarioExtenso):
        self.dataDiarioExtenso = dataDiarioExtenso
                
    def get_Data_Diario_Extenso(self):
        return self.dataDiarioExtenso         
    
    def set_Complemento(self, complementoDiario):
        self.complementoDiario = complementoDiario
                
    def get_Complemento(self):
        return self.complementoDiario       

    def set_Pagina_Diario(self, paginaDiario):
        self.paginaDiario = paginaDiario
        
    def get_Pagina_Diario(self):
        return self.paginaDiario          

    def set_Secao_Diario(self, secaoDiario):
        self.secaoDiario = secaoDiario
        
    def get_Secao_Diario(self):
        return self.secaoDiario     
    
    def set_ContadorLinhasSecao(self, contadorLinhasSecao):
        self.contadorLinhasSecao = contadorLinhasSecao
        
    def get_ContadorLinhasSecao(self):
        return self.contadorLinhasSecao    
    
    def set_ContadorLinhasGeral(self, contadorLinhasGeral):
        self.contadorLinhasGeral = contadorLinhasGeral
        
    def get_ContadorLinhasGeral(self):
        return self.contadorLinhasGeral 
    
    def set_ContadorLinhasPagina(self, contadorLinhasPagina):
        self.contadorLinhasPagina = contadorLinhasPagina
        
    def get_ContadorLinhasPagina(self):
        return self.contadorLinhasPagina         

 
def getOverlap(a, b):
    return max(0, min(a[1], b[1]) - max(a[0], b[0]))    

  
def run_program(lista_aquivos_pdf=None, diretorioArquivos=None):
    import warnings
    warnings.filterwarnings(action='ignore', category=UserWarning, module='gensim')
    #from gensim import corpora, models, similarities
    import nltk
    #from nltk import regexp_tokenize
    import os  
    import openpyxl as op
    #from pathlib import Path    
    import re
    import dataExtenso
    #import enchant
    import openpyxl
    #import similaridadeTextual
    #import transformaParaRadical    
    from statistics import median
    #import buscaSimilaridades
    #import db_connection
    #dic = enchant.Dict("pt_BR_2")  


    
    #print(lista_aquivos_txt)
    
    #for each in lista_aquivos_pdf:
        #print(each)
    
    stopwords = nltk.corpus.stopwords.words('portuguese')
    more_stopwords = "- [ ] . , ; : | ) ( i j ' / \ l s ix t � ^ ~ - # r n" # palavras sozinhas contendo esses caracteres nao irao compor a lista de palavras, serao ignoradas: sao "palavras" que nada acrescentam
    stopwords += more_stopwords.split()
    
    #titulos_blocos = "EXTRATO EXTRATOS AVISO AVISOS RATIFICAÇÃO RATIFICAÇÕES EDITAL RESULTADO PREGÃO COMUNICADO RECONHECIMENTO TERMO CHAMAMENTO" #LEILÃO DEMONSTRATIVO
    #titulos_blocos = "EXTRATO EXTRATOS AVISO AVISOS RATIFICAÇÃO RATIFICAÇÕES EDITAL JULGAMENTO PREGÃO COMUNICADO RECONHECIMENTO TERMO CHAMAMENTO LEILÃO PUBLICIDADE CONTRATO APOSTILA" #LEILÃO DEMONSTRATIVO
    #titulos_blocos = "EXTRATO EXTRATOS AVISO AVISOS RATIFICAÇÃO RATIFICAÇÕES RETIFICAÇÃO RETIFICAÇÕES EDITAL JULGAMENTO PREGÃO COMUNICADO RECONHECIMENTO TERMO CHAMAMENTO LEILÃO PUBLICIDADE CONTRATO APOSTILA ^CONCURSO" #LEILÃO DEMONSTRATIVO    
  
    
    #print(lista_titulos_blocos)
    
    #titulos_Subsecretaria = "^BRB ^ARQUIVO ^SUBSECRETARIA ^SOCIEDADE ^COMPANHIA ^CONSELHO ^AG[ÊE]NCIA ^POL[ÍI]CIA"
    titulos_Subsecretaria = "^BRB ^ARQUIVO ^SUBSECRETARIA ^SOCIEDADE ^COMPANHIA ^CONSELHO ^POL[ÍI]CIA"
    lista_Subsecretaria = titulos_Subsecretaria.split()
    lista_Subsecretaria.append('^AG[EÊ]NCIA DE FISCALIZA[ÇC][AÃ]O')  
    lista_Subsecretaria.append('^AG[EÊ]NCIA REGULADORA')      
    lista_Subsecretaria.append('^BANCO DE BRAS[ÍI]LIA')  
    lista_Subsecretaria.append('^BANCO DE BRAS[ÍI]LIA')        
    lista_Subsecretaria.append('^ADMINISTRA[ÇC][ÃA]O REGIONAL')
    lista_Subsecretaria.append('^CORPO DE BOMBEIROS MILITAR')
    lista_Subsecretaria.append('^SUPERINTEND[ÊE]NCIA DA REGI[ÃA]O DE SA[ÚU]DE')
    lista_Subsecretaria.append('^HOSPITAL ')    
    lista_Subsecretaria.append('^COMPLEXO REGULADOR EM SA[ÚU]')        
    lista_Subsecretaria.append('^POL[ÍI]CIA CIVIL DO DISTRITO FEDERAL')   
    lista_Subsecretaria.append('^EMPRESA DE')
    lista_Subsecretaria.append('^C[ÂA]MARA LEGISLATIVA DO DISTRITO FEDERAL')
    lista_Subsecretaria.append('^CASA MILITAR')
    lista_Subsecretaria.append('^FUNDO')
    lista_Subsecretaria.append('^FUNDAÇÃO')
    lista_Subsecretaria.append('^PROFLORA')    
    lista_Subsecretaria.append('^CENTRAIS DE ABASTECIMENTO')
    lista_Subsecretaria.append('^SERVIÇO DE LIMPEZA URBANA')    
    lista_Subsecretaria.append('^DEPARTAMENTO DE ESTRADAS DE RODAGEM')    
    lista_Subsecretaria.append('^TRANSPORTE URBANO DO DISTRITO FEDERAL')    
    lista_Subsecretaria.append('^SECRETARIA ADJUNTA DE MOBILI[ÁA]RIO')
    lista_Subsecretaria.append('^CENTRAIS DE ABASTECIMENTO DO DISTRITO FEDERAL')
    lista_Subsecretaria.append('^CENTRAL DE ABASTECIMENTO DE BRASÍLIA')    
    lista_Subsecretaria.append('^EMPRESA DE ASSIST[ÊE]NCIA T[ÉE]CNICA E EXTENS[A]ÃO RURAL')
    lista_Subsecretaria.append('^VENDA DIRETA DE IM[ÓO]VEIS')
    lista_Subsecretaria.append('^INSTITUTO DE PREVID[ÊE]NCIA DOS SERVIDORES')
    lista_Subsecretaria.append('^INSTITUTO DE PREVIDÊNCIA DOS SERVIDORES')
    lista_Subsecretaria.append('^INSTITUTO DO MEIO AMBIENTE')
    lista_Subsecretaria.append('^INSTITUTO DE DEFESA DO CONSUMIDOR')
    lista_Subsecretaria.append('ADJUNTA')       
    lista_Subsecretaria.append('CORREGEDORIA')
    lista_Subsecretaria.append('SUBCONTROLADORIA')
    lista_Subsecretaria.append('SECRETARIA DE CONTABILIDADE, ORÇAMENTO E FINANÇAS')
    lista_Subsecretaria.append('SECRETARIA[- ]GERAL')
    lista_Subsecretaria.append('^CONSELHO DE ADMINISTRA[ÇC][ÃA]O')

    lista_Subsecretaria.append('^SUBSECRET[ÁA]RIA DE ADMINISTRA[ÇC][ÃA]O')        
    lista_Subsecretaria.append('^PROGRAMA DE APOIO ')          

    
    #lista_Subsecretaria.append('ADJUNTA')
    
    #CENTRAIS DE ABASTECIMENTO DO DISTRITO FEDERAL
    #EMPRESA DE ASSISTÊ[E]NCIA TÉ[E]CNICA E EXTENS[A]ÃO RURAL

    
    #print(lista_Subsecretaria)
    
    lista_Secretarias = '^PODER|^CASA CIVIL|^GOVERNADORIA|^SECRETARIA|^PROCURADORIA|^DEFENSORIA|CONTROLADORIA GERAL|^TRIBUNAL DE CONTAS|^INEDITORIAIS'
    #'^PODER|^GOVERNADORIA|^SECRETARIA|^PROCURADORIA|^DEFENSORIA|ONTROLADORIA GERAL DO DISTRITO FEDERAL|^TRIBUNAL DE CONTAS|^INEDITORIAIS'
    lista_Titulares = '^A SUBSECRETARIA|^A SUBSECRET[ÁA]RIA|^O SUBSECRET[ÁA]RIO|^A SECRETARIA|^A SECRET[ÁA]RIA|^O SECRET[ÁA]RIO|^O SECRET[ÁA]RIO|^O DIRETOR|^A DIRETORA|^O GERENTE|^A GERENTE|^O PRESIDENTE|^A PRESIDENTE|^O CHEFE|^A CHEFE|^A COMISS[ÃA]O|^O FUNDO|^O CONSELHO'
    #print(lista_Subsecretaria)

    titulos_Subsecretaria_Subordinados = "^GER[ÊE]NCIA ^SUPERINTEND[ÊE]NCIA ^COORDENA[ÇC][ÃA]O ^COMISS[ÃA]O ^DEPARTAMENTO ^DIRETORIA ^ASSESSORIA ^N[ÚU]CLEO"
    #titulos_Subsecretaria_Subordinados = "^GERÊNCIA ^SUPERINTENDÊNCIA ^COORDENAÇÃO ^COMISSÃO ^DEPARTAMENTO ^DIRETORIA"    
    #titulos_Subsecretaria_Subordinados = "^GERÊNCIA ^SUPERINTENDÊNCIA ^COORDENAÇÃO ^COMISSÃO ^DEPARTAMENTO ^DIRETORIA S.A."    
    lista_Subsecretaria_Subordinados = titulos_Subsecretaria_Subordinados.split()
    lista_Subsecretaria_Subordinados.append('^CORPO DE BOMBEIROS MILITAR')
    lista_Subsecretaria_Subordinados.append('^COMIT[ÊE]')
    #lista_Subsecretaria_Subordinados.append('^SERVIÇO')
    lista_Subsecretaria_Subordinados.append('^JUNTA ADMINISTRATIVA')
    lista_Subsecretaria_Subordinados.append('^UNIDADE GESTORA DE FUNDOS')
    lista_Subsecretaria_Subordinados.append('^CONSELHO DE GEST[ÃA]O')    
    #lista_Subsecretaria_Subordinados.append('^SERVIÇO ')        
    lista_Subsecretaria_Subordinados.append('^AG[ÊE]NCIA DE DESENVOLVIMENTO')        


    #pessoas = 'Preg.* Liqui.* ecret.* Pres.*'
    #lista_pessoas = pessoas.split()
    #print(lista_palavras)
    #x = '|'.join(lista_palavras)
    #string_pagina = ""

    

    if lista_aquivos_pdf == None:
        #print("------------------------------- None ------------------------------------")                                                                       
        lista_aquivos_pdf = []                                                                                                            
        for root, _, files in os.walk(diretorioArquivos):
            for name in files:
                if name.endswith(".xlsx"):                             
                    lista_aquivos_pdf.append([os.path.join(root, name), name])              
        #    del dirs[:] # interage apenas com a primeira pasta
    
    
    #print(lista_aquivos_pdf)
    
    #if not isinstance(lista_aquivos_pdf, list):              ## Nice way to check if it's a list
        #lista_aquivos_txt = [lista_aquivos_txt]
        #lista_aquivos_pdf = [lista_aquivos_pdf]    

    #connection = db_connection.get_connection()
        
    for txts in lista_aquivos_pdf:
        #print(txts)
        pular = 0
        
        txts[0] = txts[0].replace('.pdf','.xlsx')
        txts[1] = txts[1].replace('.pdf','.xlsx')
        
        nome_Secretaria = ""
    
        titulos_blocos = []
        lista_titulos_blocos = []     
        lista_final = []
        lista_info_blocos = []

        
        print(txts[1])
        Diario = classe_Diario()
        Diario.set_Numero_Diario('')
        Diario.set_Data_Diario('')
        Diario.set_Pagina_Diario('')        
        Diario.set_Secao_Diario('')
        Diario.set_Pagina_Diario(1)
        
        if re.findall('DODF.*-', txts[1]):
            if re.findall('\d+', txts[1]):
                Diario.set_Numero_Diario(re.findall('\d+', txts[1])[0])
            if re.findall('\d\d[-.\\\\]\d\d[-.\\\\]\d\d\d?\d?', txts[1]):
                Diario.set_Data_Diario(re.findall('\d\d[-.\\\\]\d\d[-.\\\\]\d\d\d?\d?', txts[1])[0])                
                lista_data = re.findall('(\d\d)[-](\d\d)[-](\d\d\d\d)', '05-11-2018')[0]
                Diario.set_Data_Diario_Extenso(dataExtenso.transformaDataPorExtenso(Diario.get_Numero_Diario(), lista_data[0], lista_data[1], lista_data[2]))           
            if re.findall('\d\d[-.\\\\]\d\d[-.\\\\]\d\d\d?\d?[ ](.*).xlsx', txts[1]):
                Diario.set_Complemento(re.findall('\d\d[-.\\\\]\d\d[-.\\\\]\d\d\d?\d?[ ]?(.*).xlsx', txts[1])[0])                            
                
        #print(txts[0])
                              
        #txts = txts[0].replace('\\','/')
        
        #print(txts)
        #print(txts[0])        


        
        wb_obj = openpyxl.load_workbook(txts[0]) 
          
        sheet_obj = wb_obj.active 
        m_row = sheet_obj.max_row 
        m_columns = sheet_obj.max_column
        
        #print(m_columns)
        # Loop will print all values 
        # of first column
        list_linhas = []
        #list_linhas = []
        linha = []
        
        
        for i in range(2, m_row + 1):   
            for j in range (1, m_columns + 1):
                cell_obj = sheet_obj.cell(row = i, column = j) 
                linha.append(cell_obj.value)
            list_linhas.append(linha)
            linha = []

        #print(list_linhas)
        
        bloco = []
        bloco1_xo = []
        bloco1_xf = []
        bloco2_xo = []
        bloco2_xf = []
        blocoUnico_xo = []
        blocoUnico_xf = []
        
        
       
        #print(list_linhas[index][5])
            
        for elementos in list_linhas:
            #print(elementos[0])
            bloco.append(elementos[3])
            if "Bloco1" in elementos[9]:
                bloco1_xo.append(elementos[4]);
                bloco1_xf.append(elementos[5]);
            elif "Bloco2" in elementos[9]:
                bloco2_xo.append(elementos[4])
                bloco2_xf.append(elementos[5]);
            elif "BlocoUnico" in elementos[9]:
                blocoUnico_xo.append(elementos[4])
                blocoUnico_xf.append(elementos[5]);            
        
        
        
        def deslocamento(orientacaoBloco):
            deslocamento = 0
            #orientacaoBloco = ""    
            #orientacaoBloco = list_linhas[idx][11]
            #print(list_linhas[35][11])        
            if "Bloco1" in orientacaoBloco:
                deslocamento = median(bloco1_xo) + 0.3
                #print(deslocamento)
                #deslocamento = deslocamento + 0.2
            elif "Bloco2" in orientacaoBloco:
                deslocamento = median(bloco2_xo) + 0.3
                #deslocamento = deslocamento + 0.2
            elif "BlocoUnico" in orientacaoBloco:
                deslocamento = median(bloco1_xo) + 0.3 # melhor pegar a mediana do Bloco1
                #deslocamento = deslocamento + 0.2
            
            return deslocamento        
        
      
        
        def deslocamento(orientacaoBloco):
            deslocamento = 0
            #orientacaoBloco = ""    
            #orientacaoBloco = list_linhas[idx][11]
            #print(list_linhas[35][11])        
            if "Bloco1" in orientacaoBloco:
                deslocamento = median(bloco1_xo) + 0.3
                #print(deslocamento)
                #deslocamento = deslocamento + 0.2
            elif "Bloco2" in orientacaoBloco:
                deslocamento = median(bloco2_xo) + 0.3
                #deslocamento = deslocamento + 0.2
            elif "BlocoUnico" in orientacaoBloco:
                deslocamento = median(bloco1_xo) + 0.3 # melhor pegar a mediana do Bloco1
                #deslocamento = deslocamento + 0.2
            
            return deslocamento        
        
        def checaTextoCentralizado(idx, orientacaoBloco):
            centralizado = False
            intervaloTexto = [] 
            intervaloMedianoBloco =  []
            posicaoElementoCentralMediano = 0
            intervaloCentralizadoTexto = 0
            porcentagemInvervaloComum = 0.0
            #orientacaoBloco = ""    
            #orientacaoBloco = list_linhas[idx][11]
            #print(list_linhas[35][11])        
            if "Bloco1" in orientacaoBloco:
                intervaloTexto = [list_linhas[idx][4], list_linhas[idx][5]]
                
                tamanhoTexto = list_linhas[idx][5] - list_linhas[idx][4]
                tamanhoIntervaloMediano = max(bloco1_xf) - min(bloco1_xo) 
                
                intervaloMedianoBloco = [min(bloco1_xo), max(bloco1_xf)]
                metadeIntervaloTexto = (max(intervaloTexto) - min(intervaloTexto))/2
                posicaoElementoCentralMediano =  (max(intervaloMedianoBloco) + min(intervaloMedianoBloco))/2
                
                
                intervaloCentralizadoTexto = [posicaoElementoCentralMediano - metadeIntervaloTexto, posicaoElementoCentralMediano + metadeIntervaloTexto]
                
                denominador = (max(intervaloCentralizadoTexto)-min(intervaloCentralizadoTexto))
                if denominador == 0:
                    denominador = 0.01
                
                porcentagemInvervaloComum = getOverlap(intervaloTexto, intervaloCentralizadoTexto)/denominador
                
                if porcentagemInvervaloComum > 0.975 and tamanhoTexto/tamanhoIntervaloMediano < 0.999:
                    centralizado = True
                else:
                    centralizado = False                    

            elif "Bloco2" in orientacaoBloco:
                intervaloTexto = [list_linhas[idx][4], list_linhas[idx][5]]
                
                tamanhoTexto = list_linhas[idx][5] - list_linhas[idx][4]
                tamanhoIntervaloMediano = max(bloco2_xf) - min(bloco2_xo)                 
                
                intervaloMedianoBloco = [min(bloco2_xo), max(bloco2_xf)]
                metadeIntervaloTexto = (max(intervaloTexto) - min(intervaloTexto))/2
                posicaoElementoCentralMediano =  (max(intervaloMedianoBloco) + min(intervaloMedianoBloco))/2
                
                
                
                intervaloCentralizadoTexto = [posicaoElementoCentralMediano - metadeIntervaloTexto, posicaoElementoCentralMediano + metadeIntervaloTexto]
                
                denominador = (max(intervaloCentralizadoTexto)-min(intervaloCentralizadoTexto))
                if denominador == 0:
                    denominador = 0.01
                
                porcentagemInvervaloComum = getOverlap(intervaloTexto, intervaloCentralizadoTexto)/denominador
                #print(porcentagemInvervaloComum)                

                if porcentagemInvervaloComum > 0.975 and tamanhoTexto/tamanhoIntervaloMediano < 0.999:
                    centralizado = True
                else:
                    centralizado = False    

            elif "BlocoUnico" in orientacaoBloco:
                intervaloTexto = [list_linhas[idx][4], list_linhas[idx][5]]
                
                tamanhoTexto = list_linhas[idx][5] - list_linhas[idx][4]
                #tamanhoIntervaloMediano = median(blocoUnico_xf) - median(blocoUnico_xo)                   
                tamanhoIntervaloMediano = max(blocoUnico_xf) - min(blocoUnico_xo)  
                
                #intervaloMedianoBloco = [median(blocoUnico_xo), median(blocoUnico_xf)]
                intervaloMedianoBloco = [min(blocoUnico_xo), max(blocoUnico_xf)]
                metadeIntervaloTexto = (max(intervaloTexto) - min(intervaloTexto))/2
                posicaoElementoCentralMediano =  (max(intervaloMedianoBloco) + min(intervaloMedianoBloco))/2
                
                intervaloCentralizadoTexto = [posicaoElementoCentralMediano - metadeIntervaloTexto, posicaoElementoCentralMediano + metadeIntervaloTexto]
                
                denominador = (max(intervaloCentralizadoTexto)-min(intervaloCentralizadoTexto))
                if denominador == 0:
                    denominador = 0.01
                
                porcentagemInvervaloComum = getOverlap(intervaloTexto, intervaloCentralizadoTexto)/denominador                
                
                if porcentagemInvervaloComum > 0.975 and tamanhoTexto/tamanhoIntervaloMediano < 0.999:
                    centralizado = True
                else:
                    centralizado = False                
            
            
            if centralizado == False:    
                if porcentagemInvervaloComum > 0.95 and porcentagemInvervaloComum/(tamanhoTexto/tamanhoIntervaloMediano) > 1.2:
                    #print('--------------- repescagem ----------------------------------------------')
                    #print(list_linhas[idx][3])   
                    #print(porcentagemInvervaloComum)
                    #print(tamanhoTexto/tamanhoIntervaloMediano)   
                    #print(intervaloMedianoBloco)                  
                    centralizado = True
                elif porcentagemInvervaloComum > 0.90 and porcentagemInvervaloComum/(tamanhoTexto/tamanhoIntervaloMediano) > 2.5:
                    centralizado = True
                else:
                    centralizado = False                 
            
            #if list_linhas[idx][3].upper != list_linhas[idx][3]:
                #centralizado = False 
            '''
            if centralizado == False:                     
                print('--------------- centralizado ----------------------------------------------')
                print(list_linhas[idx][3])   
                print(porcentagemInvervaloComum)
                print(tamanhoTexto/tamanhoIntervaloMediano)   
                print(intervaloMedianoBloco)                        
            
            '''
            #print(list_linhas[idx][3])
            #print(porcentagemInvervaloComum)   
            '''
            print(tamanhoTexto)
            print(list_linhas[idx][5])
            print(list_linhas[idx][4])
            
            print(median(bloco1_xf))
            print(median(bloco1_xo))
            print(tamanhoIntervaloMediano)
            
            print(intervaloMedianoBloco)
            print(metadeIntervaloTexto)
            print(posicaoElementoCentralMediano)
            
            print(intervaloCentralizadoTexto)
            
            
            print(tamanhoTexto/tamanhoIntervaloMediano) 
            '''
            
            return centralizado       
        
        #for each in texto:
            #print(each)        
        
        #from unicodedata import normalize 
        #data = normalize('NFKD', data).encode('ASCII','ignore').decode('ASCII') #remove todos os acentos e cedeilhas, etc.
        Sec = classe_Secretaria('')
        Sec.set_Secretaria('')
        Sec.set_Subordinada1('')
        Sec.set_Subordinada2('')
        Sec.set_Subordinada3('')
        Bloc = classe_Bloco('')
        Bloc.set_Titulo_Bloco('')
        Bloc.set_Titulo_Subordinado1('')
        Bloc.set_Titulo_Subordinado2('')
        Bloc.set_Inicio_Bloco('')
        Bloc.set_Fim_Bloco('')
        Bloc.set_Texto_Bloco('')    
        regraFlexivel = False

        def setSecretaria(secretaria):
            if (len(str(secretaria.strip())) > 3):
                Sec.set_Secretaria(secretaria)            
            Sec.set_Contador_LinhasSecretaria(Sec.get_Contador_LinhasSecretaria() + 1)
            Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1);        
            print("Outros: " + secretaria)            
         
        def setSecSubordinada1(secretaria):
            if (len(str(secretaria.strip())) > 3):
                Sec.set_Subordinada1(secretaria)
            Sec.set_Contador_LinhasSecSubordinada1(Sec.get_Contador_LinhasSecSubordinada1() + 1) 
            Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1);              
            #print("entrei aqui")
            #print(secretaria)
            #print(setSecSubordinada1.__name__)
            print("SecSubordinada1: " + secretaria + " incluída com sucesso!")
            print("Deseja adicionar " + secretaria + " ao dicionário como Secretaria Subordinada Nível 1 ?")
            
         
        def setSecSubordinada2(secretaria):
            if (len(str(secretaria.strip())) > 3):
                Sec.set_Subordinada2(secretaria)
            Sec.set_Contador_LinhasSecSubordinada2(Sec.get_Contador_LinhasSecSubordinada2() + 1)  
            Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1);      
            print("Outros: " + secretaria)     
        
        def setSecSubordinada3(secretaria):
            if (len(str(secretaria.strip())) > 3):
                Sec.set_Subordinada3(secretaria)
            Sec.set_Contador_LinhasSecSubordinada3(Sec.get_Contador_LinhasSecSubordinada3() + 1) 
            Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1);   
            print("Outros: " + secretaria) 
                     
        switcher = {
                0: setSecretaria,
                1: setSecSubordinada1,
                2: setSecSubordinada2,
                3: setSecSubordinada3
            }
        
         
        def corrigeSecretariaFaltante(argument, secretaria):
            # Get the function from switcher dictionary
            func = switcher.get(argument, "nothing")
            # Execute the function
            return func(secretaria)


                        
        def corrigeModeloBloco(i, index):

                if list_linhas[index + i][12] - Diario.get_ContadorLinhasGeral() == 1:
                    choices = {0: Sec.get_Secretaria(), 1: Sec.get_Subordinada1(), 2: Sec.get_Subordinada2(), 3: Sec.get_Subordinada3()}
                    
                    elemento = 0        
                    for i in range(len(choices)):
                        if(choices.get(i) == ''):
                            elemento = i
                            break                
                                
                    if elemento > 0:
                        corrigeSecretariaFaltante(elemento, list_linhas[Diario.get_ContadorLinhasGeral() - 1][3])
                    else:
                        print("Não foi possível corrigir elemento faltante. Todas as secretarias estão preenchidas!")
                else:                                    
                    #print("Divergência precisa ser corrigida manualmente")
                    print("Divergência tentará ser corrigida automaticamente (verificar o resultado)")
                    print(" > SEC: " + Sec.get_Secretaria())
                    #print(Sec.get_Contador_LinhasSecSubordinada1())
                    #print(Sec.get_Contador_LinhasSecSubordinada2())
                    #print(Sec.get_Contador_LinhasSecSubordinada3())
                    print(" > SEC1: " + Sec.get_Subordinada1())
                    #print(Sec.get_Subordinada2())
                    #print(Sec.get_Subordinada3())
                    #print("contadorGeral - 1: " + str(Diario.get_ContadorLinhasGeral() - 1))
                    print("         > BLOCO DIVERGENTE: ")
                    #print(Sec.get_Subordinada2())
                    #print(Sec.get_Subordinada3())
                    #print("contadorGeral - 1: " + str(Diario.get_ContadorLinhasGeral() - 1))
                    #print(index)
                    #print(list_linhas[Diario.get_ContadorLinhasGeral() - 1][3])
                    #print(int(list_linhas[index + i - 1][12]))
                    
                    for k in range (Diario.get_ContadorLinhasGeral() - 1, int(list_linhas[index + i - 1][12])):
                        #print(list_linhas[k][3])
                        print("         " + list_linhas[k][3])  
                        #x = checaTextoCentralizado(k, list_linhas[k][9])
                        #print(x)    
                        #x = deslocamento(list_linhas[k][9])
                        #print(x)             
                        #print(list_linhas[k][4])
                    
                    montaBloco(Diario.get_ContadorLinhasGeral() - 1, list_linhas[Diario.get_ContadorLinhasGeral() - 1][3], pular, True, int(list_linhas[index + i - 1][12])) 
                    #regraFlexivel = False


                    #Diario.set_ContadorLinhasGeral(int(list_linhas[index][12]));
                    #print("contadorGeral: " + str(Diario.get_ContadorLinhasGeral()))
                    #Diario.set_ContadorLinhasGeral(int(list_linhas[index][12]));
                #print(consultaSecretariaCorrigida(elemento))
                #if consultaSecretariaCorrigida(elemento) != '':                          

        
        def atualizaContadoresSecao():
            #Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)
            #Diario.set_ContadorLinhasPagina(Diario.get_ContadorLinhasPagina() + 1)  
            Diario.set_ContadorLinhasSecao(Diario.get_ContadorLinhasSecao() + 1)
            Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)         
            if (Diario.get_ContadorLinhasGeral() != list_linhas[index][12]):
                print('Erro no reconhecimento do elemento anterior à Seção do DODF')                
        
        def atualizaContadoresSecretaria(i, index):
            #Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)
            #Diario.set_ContadorLinhasPagina(Diario.get_ContadorLinhasPagina() + 1)
            Sec.set_Contador_LinhasSecretaria(Sec.get_Contador_LinhasSecretaria() + 1)
            Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1) 
            #print(list_linhas[index + i][12])
            #print(Diario.get_ContadorLinhasGeral())       
            if (Diario.get_ContadorLinhasGeral() != list_linhas[index + i][12]):
                #Sec.set_Contador_LinhasSecretaria(Sec.get_Contador_LinhasSecretaria() - i - 1)
                #Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() - i - 1)                 
                print('Erro no reconhecimento do elemento anterior à Secretaria')
                #print(i)
                #print(Diario.get_ContadorLinhasGeral() - i - 1)                              
                #print(list_linhas[index - i - 1][12])                
                #corrigeModeloSec(i, index)
                Sec.set_Contador_LinhasSecretaria(Sec.get_Contador_LinhasSecretaria() - 1)
                corrigeModeloBloco(i, index)                
            
        def atualizaContadoresSecSubordinada1(i, index):
            #Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)
            #Diario.set_ContadorLinhasPagina(Diario.get_ContadorLinhasPagina() + 1)
            Sec.set_Contador_LinhasSecSubordinada1(Sec.get_Contador_LinhasSecSubordinada1() + 1) 
            Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)       
            if (Diario.get_ContadorLinhasGeral() != list_linhas[index + i][12]):
                print('Erro no reconhecimento do elemento anterior à Secretaria Subordinada1')
                
                corrigeModeloBloco(i, index)               

            
        def atualizaContadoresSecSubordinada2(i, index):
            #Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)
            #Diario.set_ContadorLinhasPagina(Diario.get_ContadorLinhasPagina() + 1)
            Sec.set_Contador_LinhasSecSubordinada2(Sec.get_Contador_LinhasSecSubordinada2() + 1)  
            Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)      
            if (Diario.get_ContadorLinhasGeral() != list_linhas[index + i][12]):
                print('Erro no reconhecimento do elemento anterior à Secretaria Subordinada2')
                corrigeModeloBloco(i, index)   
                

        def atualizaContadoresSecSubordinada3(i, index):
            #Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)
            #Diario.set_ContadorLinhasPagina(Diario.get_ContadorLinhasPagina() + 1)
            Sec.set_Contador_LinhasSecSubordinada3(Sec.get_Contador_LinhasSecSubordinada3() + 1) 
            Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)          
            if (Diario.get_ContadorLinhasGeral() != list_linhas[index + i][12]):
                print('Erro no reconhecimento do elemento anterior à Secretaria Subordinada3')
                corrigeModeloBloco(i, index)
            
        def atualizaContadoresTitBloco(i, index, Bloc):
            #Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)
            #Diario.set_ContadorLinhasPagina(Diario.get_ContadorLinhasPagina() + 1)            
            #print(str(Bloc.get_Contador_LinhasSubTitBloco()) + " atual")
            Bloc.set_Contador_LinhasTitBloco(Bloc.get_Contador_LinhasTitBloco() + 1)
            Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)
            #print(str(Bloc.get_Contador_LinhasSubTitBloco()) + " atual")                   
            #print('------------------------ passei aqui ------------------- ')                
            if (Diario.get_ContadorLinhasGeral() != list_linhas[index][12]):
                print('Erro no reconhecimento do elemento anterior ao Titulo de Bloco: ' + Bloc.get_Titulo_Bloco())             
                corrigeModeloBloco(i, index)

        def atualizaContadoresSubTitBloco(i, index, Bloc):
            #Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)
            #Diario.set_ContadorLinhasPagina(Diario.get_ContadorLinhasPagina() + 1)

            Bloc.set_Contador_LinhasSubTitBloco(Bloc.get_Contador_LinhasSubTitBloco() + 1)
            Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)  
       
            if (Diario.get_ContadorLinhasGeral() != list_linhas[index + i][12]):
                print('Erro no reconhecimento do elemento anterior ao Subtítulo do Bloco')                          
                corrigeModeloBloco(i, index)

        def atualizaContadoresDemaisSubTitBloco(i, index, Bloc):
            #Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)
            #Diario.set_ContadorLinhasPagina(Diario.get_ContadorLinhasPagina() + 1)
            Bloc.set_Contador_LinhasDemaisSubTitBloco(Bloc.get_Contador_LinhasDemaisSubTitBloco() + 1)
            Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)               
            if (Diario.get_ContadorLinhasGeral() != list_linhas[index + i][12]):
                print('Erro no reconhecimento do elemento anterior aos Demais Subtítulos do Bloco')             
                corrigeModeloBloco(i, index) 
                    
        def montaBloco(index, each_line, pular, tipoBusca, ultimoElemento):
            regraFlexivel = tipoBusca
            stop = ultimoElemento
            quebraLight = False 
            if (re.compile('|'.join(lista_titulos_blocos)).search(each_line) and list_linhas[index][4] > deslocamento(list_linhas[index][9]) and checaTextoCentralizado(index, list_linhas[index][9]) is True and not re.compile(lista_Secretarias).search(each_line) and not re.compile('|'.join(lista_Subsecretaria)).search(each_line) and not re.compile('|'.join(lista_Subsecretaria_Subordinados)).search(each_line) and not re.compile('^PROCESSO').search(each_line) or regraFlexivel): #re.IGNORECASE is used to ignore case                    
                Bloc = classe_Bloco(str(each_line))
                Diario.set_Pagina_Diario(list_linhas[index][2])
                i = 0
                w = 0
                textoBloco = ""
                #print(str(Bloc.get_Contador_LinhasTitBloco()) + " loop")
                if (tipoBusca == False):                 
                    atualizaContadoresTitBloco(i, index, Bloc)
                else:
                    Bloc.set_Contador_LinhasTitBloco(Bloc.get_Contador_LinhasTitBloco() + 1) 
                    Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)
                    #quebraLight = False                  
                    #print(index)    
                    #print(each_line)
                    #print(tipoBusca)  
                    #print(ultimoElemento)
                    #print(pular)                                     

                while (index + i <= stop):
                    i = i + 1
                    if len(str(bloco[index + i]).strip()) > 0:
                        if str(bloco[index + i]).upper() == str(bloco[index + i]) and list_linhas[index + i][4] > deslocamento(list_linhas[index + i][9]) and checaTextoCentralizado(index + i, list_linhas[index + i][9]) and re.compile(lista_Secretarias).search(str(bloco[index + i])) is None and re.compile('|'.join(lista_Subsecretaria)).search(str(bloco[index + i])) is None and re.compile('|'.join(lista_Subsecretaria_Subordinados)).search(str(bloco[index + i])) is None and re.compile('^PROCESSO').search(str(bloco[index + i])) is None and re.compile(lista_Titulares).search(str(bloco[index + i])) is None:
                            Bloc.set_Titulo_Subordinado1(str(bloco[index + i]))
                            #print(Bloc.get_Contador_LinhasTitBloco())
                            if (tipoBusca == False):                    
                                atualizaContadoresSubTitBloco(i, index, Bloc) 
                            else:                                
                                Bloc.set_Contador_LinhasTitBloco(Bloc.get_Contador_LinhasTitBloco() + 1) 
                                Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)                                                        
                            nome_Titulo_Bloco_Subordinado2 = ""
                            i = i + 1
                            #print(Bloc.get_Contador_LinhasTitBloco())
                            while len(str(bloco[index + i]).strip()) > 0 and list_linhas[index + i][4] > deslocamento(list_linhas[index + i][9]) and checaTextoCentralizado(index + i, list_linhas[index + i][9]) and str(bloco[index + i]).upper() == str(bloco[index + i]) and (re.compile(lista_Secretarias).search(str(bloco[index + i])) is None or list_linhas[index + i][8] < 12) and re.compile('|'.join(lista_Subsecretaria)).search(str(bloco[index + i])) is None and re.compile('|'.join(lista_Subsecretaria_Subordinados)).search(str(bloco[index + i])) is None and re.compile('^PROCESSO').search(str(bloco[index + i])) is None and re.compile(lista_Titulares).search(str(bloco[index + i])) is None:
                                nome_Titulo_Bloco_Subordinado2 = nome_Titulo_Bloco_Subordinado2 + " " + str(bloco[index + i]) + "\n"  
                                if (tipoBusca == False):                    
                                    atualizaContadoresDemaisSubTitBloco(i, index, Bloc)
                                else:
                                    Bloc.set_Contador_LinhasTitBloco(Bloc.get_Contador_LinhasTitBloco() + 1) 
                                    Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + 1)                                                                   
                                i = i + 1
                            i = i - 1
                            if nome_Titulo_Bloco_Subordinado2 != '':                          
                                Bloc.set_Titulo_Subordinado2(nome_Titulo_Bloco_Subordinado2)                                                                                            
                            #pular = pular + i
                                                                                                                                          
                        else:                       
                            Bloc.set_Inicio_Bloco(str(bloco[index + i]))
                            Diario.set_Pagina_Diario(list_linhas[index + i][2])
                            Bloc.set_Pagina_Inicio_Bloco(Diario.get_Pagina_Diario())
                            Bloc.set_Linha_Inicio_Bloco(list_linhas[index + i][10])
                                             
                            if (tipoBusca == False and Diario.get_ContadorLinhasGeral() + 1 != list_linhas[index + i][12]):
                                print(Diario.get_ContadorLinhasGeral() + 1)
                                print(list_linhas[index + i][12])
                                print('problemas')
            
                            
                            w = i
                            while (index + w < stop):
                                if re.compile('^INEDIT.*').search(str(bloco[index + w]).replace(' ', '')) and list_linhas[index + w][8] > 11:   
                                    pular = pular + stop - index
                                    break
                                
                                if re.compile('^SE[ÇC][ÃA]O II[I1l]$|^SEÇÃO III').search(str(bloco[index + w])) and list_linhas[index + w][8] > 11:   
                                    w = w - 1
                                    break       
                                

                                if list_linhas[index +  w][8] == 0 and list_linhas[index +  w][3] == '.':
                                    quebraLight = True    
                                    #print("quebra leve")                     
                                    
                                if list_linhas[index + w][2] > Diario.get_Pagina_Diario():
                                    print('pagina: ' + str(Diario.get_Pagina_Diario()))
                                    Diario.set_Pagina_Diario(list_linhas[index + w][2]) 
                                    Diario.set_ContadorLinhasPagina(0)
                                
                                if str(bloco[index + w]).upper() == str(bloco[index + w]) and (re.compile('|'.join(lista_titulos_blocos)).search(str(bloco[index + w])) and list_linhas[index + w][4] > deslocamento(list_linhas[index + w][9])) and not re.compile('^PROCESSO').search(str(bloco[index + w])) or (re.compile(lista_Secretarias).search(str(bloco[index + w])) and list_linhas[index + w][8] > 11) or re.compile('|'.join(lista_Subsecretaria)).search(str(bloco[index + w]))  or re.compile('|'.join(lista_Subsecretaria_Subordinados)).search(str(bloco[index + w])) :                                    
                                    
                                    
                                    if Sec.get_Secretaria != '' and re.compile(Sec.get_Secretaria().replace("(*)","")).search(str(bloco[index + w])):
                                        if len(str(bloco[index + w]).strip()) > 0:
                                            textoBloco = textoBloco + str(bloco[index + w]) + " " + "\n"                                        
                                        w = w + 1                                         
                                    elif Sec.get_Subordinada1() != '' and re.compile(Sec.get_Subordinada1().replace("(*)","")).search(str(bloco[index + w])):
                                        if len(str(bloco[index + w]).strip()) > 0:
                                            textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                        w = w + 1   
            
                                    elif Sec.get_Subordinada2() != '' and re.compile(Sec.get_Subordinada2().replace("(*)","")).search(str(bloco[index + w])):
                                        if len(str(bloco[index + w]).strip()) > 0:
                                            textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                        w = w + 1           
                                    elif Sec.get_Subordinada3() != '' and re.compile(Sec.get_Subordinada3().replace("(*)","")).search(str(bloco[index + w])):
                                        if len(str(bloco[index + w]).strip()) > 0:
                                            textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                        w = w + 1                                                                                 
                                    elif textoBloco.count('\n') + 1 < 4:
                                        if len(str(bloco[index + w]).strip()) > 0:
                                            textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                        w = w + 1
                                    elif (re.compile('|'.join(lista_titulos_blocos)).search(str(bloco[index + w])) and list_linhas[index + w][4] > deslocamento(list_linhas[index + w][9]) is False or checaTextoCentralizado(index + w, list_linhas[index + w][9]) is False):                                            
                                        if len(str(bloco[index + w]).strip()) > 0:
                                            textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                        w = w + 1                                                                                                                                                                                 
                                    elif list_linhas[index + w][4] > deslocamento(list_linhas[index + w][9]) is False or checaTextoCentralizado(index + w, list_linhas[index + w][9]) is False:                                                                                
                                        if len(str(bloco[index + w]).strip()) > 0:
                                            textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                        w = w + 1     
                                                                                                                                                                                                                                                                               
                                    else:
                                        if str(bloco[index + w]).upper() != str(bloco[index + w]):      
                                            if len(str(bloco[index + w]).strip()) > 0:
                                                textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                            w = w + 1                                             
                                        elif checaTextoCentralizado(index + w, list_linhas[index + w][9]) is False:
                                            if len(str(bloco[index + w]).strip()) > 0:
                                                textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                            w = w + 1
                                        elif len(str(bloco[index + w]).strip()) < 2:
                                            if len(str(bloco[index + w]).strip()) > 0:
                                                textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                            w = w + 1                                            
                                        elif textoBloco.count('\n') + 1 < 4:
                                            if len(str(bloco[index + w]).strip()) > 0:
                                                textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                            w = w + 1                                            
                                                        
                                        elif re.compile(lista_Secretarias).search(str(bloco[index + w])):
                                            #print('str(bloco[index + w]) :' + str(bloco[index + w]))
                                            if list_linhas[index + w][8] < 11 and "bold" not in str(list_linhas[index + w][8]).lower():
                                                if len(str(bloco[index + w]).strip()) > 0:
                                                    textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                                w = w + 1
                                            else:
                                                w = w -1                                                        
                                                break 
                      
                                        elif re.compile('|'.join(lista_Subsecretaria)).search(str(bloco[index + w])):
                                            #print('str(bloco[index + w]) :' + str(bloco[index + w]))
                                            if list_linhas[index + w][8] < 9 and "bold" not in str(list_linhas[index + w][8]).lower():
                                                if len(str(bloco[index + w]).strip()) > 0:
                                                    textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                                w = w + 1
                                            else:
                                                w = w -1                                                        
                                                break                                             
                                           
                                        else:
                                            if str(bloco[index + w]).upper() == str(bloco[index + w]): 
          
          
                                                if  quebraLight == True:
                                                    loop = index + w
                                                    while (loop > index + i):
                                                        #print(loop)
                                                        #print(list_linhas[loop][8])
                                                        #print(list_linhas[loop][3])
                                                        if list_linhas[loop][8] == 0 and list_linhas[loop][3] == '.':
                                                            #print("achou1")
                                                            if list_linhas[loop][6] == list_linhas[index + w][6]:
                                                                #print("achou2")
                                                                if len(str(bloco[index + w]).strip()) > 0:
                                                                    textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                                                w = w + 1
                                                                #print("localizou elemento na tabela com inicio no mesmo Y")
                                                                break                                                                       
                                                                #textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                                                #w = w + 1                                                                               
                                                            #    print("mesmo Y")
                                                            #textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                                            #w = w + 1  
                                                        loop = loop - 1
                                                    if (loop == index + i):
                                                        if re.compile('|'.join(lista_titulos_blocos)).search(str(bloco[index + w])):                                                     
                                                            if (index + w - 1 > index + i):                                                        
                                                                if re.compile("^ANEXO I|^ANEXO II").search(str(bloco[index + w - 1])): 
                                                                    #print("eu aqui2")
                                                                    #print(str(bloco[index + w - 1]))
                                                                    #print(str(bloco[index + w]))
                                                                    textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                                                    w = w + 1                                                           
                                                                else:
                                                                    w = w - 1
                                                                    break                                                                   
                                                            else:
                                                                w = w - 1
                                                                break                                                           
                                                        
                                                        else:
                                                            w = w -1                                                        
                                                            break
                                                else:                                                                      
                                                                                                                            
                                                    #if re.compile("\.\.\.|;|--|COMISSÃO ESPECIAL DE LICITAÇÃO - PEDIDO DE PROPOSTA Nº.03/2019-SEDF - CPL").search(str(bloco[index + w])):
                                                    if re.compile("\.\.\.|;").search(str(bloco[index + w])):                                                        
                                                        if len(str(bloco[index + w]).strip()) > 0:
                                                            textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                                        w = w + 1                                                                                                     
                                                    
                                                    elif re.compile('|'.join(lista_titulos_blocos)).search(str(bloco[index + w])):                                                     
                                                        if (index + w - 1 > index + i):                                                        
                                                            if re.compile("^ANEXO I|^ANEXO II").search(str(bloco[index + w - 1])): 
                                                                #print("eu aqui1")
                                                                #print(str(bloco[index + w - 1]))
                                                                #print(str(bloco[index + w]))
                                                                textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                                                w = w + 1                                                           
                                                            else:
                                                                w = w - 1
                                                                break                                                                   
                                                        else:
                                                            w = w - 1
                                                            break                                                     
                                                                                                        
                                                    elif (str(bloco[index + w - 1]).upper() != str(bloco[index + w - 1]) and list_linhas[index +  w][6] - list_linhas[index + w - 1][6] < 10):
                                                        if (index + w + 1 < stop):
                                                            if (str(bloco[index + w + 1]).upper() != str(bloco[index + w + 1]) and list_linhas[index +  w + 1][6] - list_linhas[index + w][6] < 10):
                                                                textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                                                w = w + 1     
                                                            elif list_linhas[index +  w + 1][8] == 0 and  list_linhas[index +  w + 1][3] == '.':
                                                                textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                                                w = w + 1                                                                                                                        
                                                                                                                                             
                                                                
                                                            else:
                                                                                                                                   
                                                                #print("else1")
                                                                #print(str(bloco[index + w]))
                                                                #print(checaTextoCentralizado(index + w, list_linhas[index + w][9]))
                                                                
                                                                w = w - 1
                                                                break                                                                   
                                                        else:
                                                            #print("else2")
                                                            #print(str(bloco[index + w]))
                                                            #print(checaTextoCentralizado(index + w, list_linhas[index + w][9]))
                                                            w = w - 1
                                                            break                                                        
                                                                                                                                                        
                                                    
                                                    else:                                           
                                                    #print(str(bloco[index + w]))
                                                    #print(checaTextoCentralizado(index + w, list_linhas[index + w][9]))
                                                        #print("else3")
                                                        w = w - 1
                                                        break                                                    
                                                    #re.compile('|'.join(lista_titulos_blocos)).search(str(bloco[index + w]) 
                                                            
                                                    #elif re.compile("^ANEXO I|^ANEXO II").search(str(bloco[index + w])):
                                                    #    print('str(bloco[index + w])' + str(bloco[index + w]))
                                                    #    if len(str(bloco[index + w]).strip()) > 0:
                                                    #        textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                                    #    w = w + 1                                                                            
                                                    ''' (antes do else de cima)
                                                    elif (str(bloco[index + w - 1]).upper() != str(bloco[index + w - 1]) and list_linhas[index +  w][6] - list_linhas[index + w - 1][6] < 10):
                                                        if (index + w + 1 < stop):
                                                            if (str(bloco[index + w + 1]).upper() != str(bloco[index + w + 1]) and list_linhas[index +  w + 1][6] - list_linhas[index + w][6] < 10):
                                                                textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                                                w = w + 1     
                                                            elif list_linhas[index +  w + 1][8] == 0 and  list_linhas[index +  w + 1][3] == '.':
                                                                textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                                                w = w + 1                                                                                                                        
                                                                                                                                             
                                                                
                                                            else:
                                                                                                                                   
                                                                #print("else1")
                                                                #print(str(bloco[index + w]))
                                                                #print(checaTextoCentralizado(index + w, list_linhas[index + w][9]))
                                                                
                                                                w = w - 1
                                                                break                                                                   
                                                        else:
                                                            #print("else2")
                                                            #print(str(bloco[index + w]))
                                                            #print(checaTextoCentralizado(index + w, list_linhas[index + w][9]))
                                                            w = w - 1
                                                            break                                                        
                                                
                                                    '''  

                                                                                    
                                            else:
                                                textoBloco = textoBloco + str(bloco[index + w])  + " " + "\n"                                        
                                                w = w + 1                                                                                                                                                                                                                                     
                                else:
                                    if len(str(bloco[index + w]).strip()) > 0:
                                        textoBloco = textoBloco + str(bloco[index + w]) + "\n"                                                                                
                                    w = w + 1   
                          
                                                                                                                               
                            break
                if w == 0:
                    Diario.set_Pagina_Diario(list_linhas[index + i - 1][2])
                else:
                    Diario.set_Pagina_Diario(list_linhas[index + w - 1][2])
                
                listaLinhas = []
                textoBlocoDefinitivo = ''
                analiseTextoBloco = re.split('\n', textoBloco)
                for indice, line in enumerate(analiseTextoBloco):
                    line = line.replace(os.linesep,'')                            
                    if len(line.strip()) > 0: #and not  == len(analiseTextoBloco) - 1:
                        if len(listaLinhas) > 0 and len(listaLinhas) < 2 and len(line.strip()) < 2:   
                            listaLinhas[len(listaLinhas) - 1] = listaLinhas[len(listaLinhas) - 1] + line
                        else:         
                            listaLinhas.append(line)                                                                    
                                            
                textoBlocoDefinitivo = ""
                for indice, lines in enumerate(listaLinhas):
                    if indice == len(listaLinhas) - 1:
                        textoBlocoDefinitivo = textoBlocoDefinitivo + lines    
                    else:
                        textoBlocoDefinitivo = textoBlocoDefinitivo + lines + '\n'                                
                
                   
                contadorLinhas = textoBlocoDefinitivo.count('\n') + 1
                ultimaLinha = textoBlocoDefinitivo.split('\n')[contadorLinhas - 1]
                #print(ultimaLinha)
            
                Bloc.set_Texto_Bloco(textoBlocoDefinitivo)
                Bloc.set_Pagina_Fim_Bloco(Diario.get_Pagina_Diario())
                
                if index + w >= len(list_linhas):
                    Bloc.set_Linha_Fim_Bloco(list_linhas[index + w - 1][10])
                else:
                    Bloc.set_Linha_Fim_Bloco(list_linhas[index + w][10])

                Bloc.set_Contador_Linhas(contadorLinhas)
                Bloc.set_Fim_Bloco(ultimaLinha)
                
                Bloc.set_Contador_LinhasTotalBloco(Diario.get_ContadorLinhasSecao() + Bloc.get_Contador_Linhas() + Sec.get_Contador_LinhasSecretaria() + Sec.get_Contador_LinhasSecSubordinada1() + Sec.get_Contador_LinhasSecSubordinada2() + Sec.get_Contador_LinhasSecSubordinada3() + 
                                                   Bloc.get_Contador_LinhasTitBloco() + Bloc.get_Contador_LinhasSubTitBloco() + Bloc.get_Contador_LinhasDemaisSubTitBloco())
                
                Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + Bloc.get_Contador_Linhas())
                                     
                pular = pular + w
                    
                lista_info_blocos = [Diario.get_Numero_Diario(), Diario.get_Data_Diario(), Diario.get_Complemento(), Diario.get_Secao_Diario(), 
                                     Sec.get_Secretaria(), Sec.get_Subordinada1(), Sec.get_Subordinada2(), Sec.get_Subordinada3(), 
                                     Bloc.get_Titulo_Bloco(), Bloc.get_Titulo_Subordinado1(), Bloc.get_Titulo_Subordinado2(), Bloc.get_Texto_Bloco(), Bloc.get_Inicio_Bloco(), Bloc.get_Fim_Bloco(), Bloc.get_Pagina_Inicio_Bloco(),  Bloc.get_Linha_Inicio_Bloco(), Bloc.get_Pagina_Fim_Bloco(), Bloc.get_Linha_Fim_Bloco(), Bloc.get_Contador_Linhas(), 
                                     Diario.get_ContadorLinhasSecao(), Sec.get_Contador_LinhasSecretaria(), Sec.get_Contador_LinhasSecSubordinada1(), Sec.get_Contador_LinhasSecSubordinada2(), Sec.get_Contador_LinhasSecSubordinada3(), 
                                     Bloc.get_Contador_LinhasTitBloco(), Bloc.get_Contador_LinhasSubTitBloco(), Bloc.get_Contador_LinhasDemaisSubTitBloco(), Bloc.get_Contador_LinhasTotalBloco()]
                lista_final.append(lista_info_blocos)
                Diario.set_ContadorLinhasSecao(0)
                Sec.set_Contador_LinhasSecretaria(0)
                Sec.set_Contador_LinhasSecSubordinada1(0)
                Sec.set_Contador_LinhasSecSubordinada2(0)
                Sec.set_Contador_LinhasSecSubordinada3(0)
                #Bloc.set_Contador_LinhasTitBloco(0)
                regraFlexivel = False    
            return pular         
        
        for index, each_line in enumerate(bloco):
            
            #each_line = elementos[0]
            
            #print(each_line)
            
            #print(bloco[index])
            #print(list_linhas[index][5])
             
            
            if re.compile('^SE[ÇC][ÃA]O I[I1l]$').search(str(each_line)) and list_linhas[index][8] > 11: 
                Diario.set_Secao_Diario('SEÇÃO II')
                print("----> SEÇÃO II") #controlePrint
                lista_titulos_blocos = []
                lista_titulos_blocos.append('^PORTARIA')      
                lista_titulos_blocos.append('^RESOLU[ÇC][ÃA]O')
                lista_titulos_blocos.append('^RESOLU[ÇC][ÕO]ES')
                lista_titulos_blocos.append('^INSTRU[ÇC][ÃA]O')
                lista_titulos_blocos.append('^INSTRU[ÇC][ÕO]ES')                
                lista_titulos_blocos.append('^ORDEM DE SERVI[ÇC]O[S]?')
                lista_titulos_blocos.append('^DECRETO[S]?')
                lista_titulos_blocos.append('^DESPACHO[S]?')
                lista_titulos_blocos.append('^ORDENS DE SERVI[ÇC]O[S]?')
                lista_titulos_blocos.append('^RETIFICA[ÇC][ÃA]O')
                lista_titulos_blocos.append('^DETERMINAÇÃO')
                lista_titulos_blocos.append('^EXTRATO D[AE]')
                lista_titulos_blocos.append('^DOCUMENTO DECISÓRIO')                                                         
                    
                lista_titulos_blocos.append('^RETIFICA[ÇC][ÃA]O')  
                lista_titulos_blocos.append('^RETIFICA[ÇC][ÕO]ES')
                lista_titulos_blocos.append('^APOSTILAMENTO[S]?')  
                
                atualizaContadoresSecao()
                
                #Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + Diario.get_ContadorLinhasSecao());  
                #if (Diario.get_ContadorLinhasGeral() != list_linhas[index][12]):
                #    print('problemas')
                    
                #contadorLinhasGeral = contadorLinhasGeral + 1          
                #contadorLinhasPagina = contadorLinhasPagina + 1
                
            if re.compile('^SE[ÇC][ÃA]O II[I1l]$|^SEÇÃO III').search(str(each_line)) and list_linhas[index][8] > 11: 
                Diario.set_Secao_Diario('SEÇÃO III')
                
                '''
                Sec = classe_Secretaria()
                Sec.set_Secretaria('')
                Sec.set_Subordinada1('')
                Sec.set_Subordinada2('')
                Sec.set_Subordinada3('')
                Bloc = classe_Bloco()
                Bloc.set_Titulo_Bloco('')
                Bloc.set_Titulo_Subordinado1('')
                Bloc.set_Titulo_Subordinado2('')
                Bloc.set_Inicio_Bloco('')
                Bloc.set_Fim_Bloco('')
                Bloc.set_Texto_Bloco('')                  
                #pular = pular + 1
                '''
                
                print("----> SEÇÃO III") #controlePrint     
                titulos_blocos = []
                lista_titulos_blocos = []
                titulos_blocos = "EXTRATO EXTRATOS AVISO AVISOS RATIFICA[ÇC][ÃA]O RATIFICA[ÇC][ÕO]ES RETIFICA[ÇC][AÃ]O RETIFICA[ÇC][ÕO]ES EDITAL JULGAMENTO PREG[ÃA]O COMUNICADO RECONHECIMENTO TERMO CHAMAMENTO LEIL[ÃA]O PUBLICIDADE CONTRATO APOSTILA ^CONCURSO ^CONCORR[ÊE]NCIA ^CONVITE ^CONVOCA[ÇC][ÃA]O ^DECLARAÇÕES ^DECLARA[ÇC][ÃA]O ^NOTIFICA[ÇC][ÃA]O ^NOTIFICA[ÇC][ÕO]ES ^REGULAMENTO ^POSICIONAMENTO" #INCLUIDO RETIFICACAO E REFIFICACEOS
                #titulos_blocos = "EXTRATO EXTRATOS AVISO AVISOS RATIFICAÇÃO RATIFICAÇÕES RETIFICAÇÃO RETIFICAÇÕES EDITAL JULGAMENTO PREGÃO COMUNICADO RECONHECIMENTO TERMO CHAMAMENTO LEILÃO PUBLICIDADE CONTRATO APOSTILA ^CONCURSO ^CONCORRÊNCIA ^CONVITE ^CONVOCAÇÃO" #INCLUIDO RETIFICACAO E REFIFICACEOS
                #titulos_blocos = "^EXTRATO ^EXTRATOS ^AVISO ^AVISOS.*REVOGAÇÃO ^RATIFICAÇÃO ^RATIFICAÇÕES ^EDITAL ^RESULTADO ^PREGÃO ^PREGAO ^COMUNICADO ^RECONHECIMENTO ^TERMO ^CHAMAMENTO ^PROCESSO"
                lista_titulos_blocos = titulos_blocos.split()
                lista_titulos_blocos.append('^APLICA[ÇC][ÃA]O DE PENALIDADE')
                lista_titulos_blocos.append('^APLICA[ÇC][ÕO]ES DE PENALIDADE')
                lista_titulos_blocos.append('^APLICA[ÇC][ÕO]ES DE PENALIDADES')        
                lista_titulos_blocos.append('^ORDEM DE SERVI[ÇC]O[S]?')
                lista_titulos_blocos.append('^LICEN[ÇC]A')                
                lista_titulos_blocos.append('^AV I S O')
                lista_titulos_blocos.append('^RECONVOCAÇÃO')
                lista_titulos_blocos.append('CARTA CONVITE')
                lista_titulos_blocos.append('^HABILITAÇÃO')                
                #lista_titulos_blocos.append('HABILITAÇÃO DE LICITANTE')         
                lista_titulos_blocos.append('^ANÁLISE[S]? DE RECURSO[S]?')
                #lista_titulos_blocos.append('RECURSO')
                lista_titulos_blocos.append('^TOMADA DE CONTAS')                
                lista_titulos_blocos.append('^DEMONSTRATIVO[S]?')                
                lista_titulos_blocos.append('^ASSEMBL[EÉ]IA[- ]GERAL')      
                lista_titulos_blocos.append('^ATA DE REGISTRO DE PREÇO')      
                 
                lista_titulos_blocos.append('^LICENCIAMENTO AMBIENTAL')   
                                                      
                lista_titulos_blocos.append('^ORDENS DE SERVI[ÇC]O[S]?')
                #lista_titulos_blocos.append('^RESULTADO FINAL|^RESULTADO PROVISÓRIO|RESULTADO DE LICITAÇÃO')  
                lista_titulos_blocos.append('^RESULTADO')          
                lista_titulos_blocos.append('^DOCUMENTO[S]? DECIS[ÓO]RIO[S]?')
                lista_titulos_blocos.append('^RELA[ÇC][ÃA]O DE COMPRAS')    
                lista_titulos_blocos.append('^RELA[ÇC][ÕO]ES DE COMPRAS') 
                lista_titulos_blocos.append('^QUADRO[S]? DEMONSTRATIVO[S]?')  
                lista_titulos_blocos.append('^RESULTADO[S]? DO CREDENCIAMENTO[S]?')
                
                
                lista_titulos_blocos.append('^CONSULTA AOS REGISTROS DE PRE[ÇC]OS')  
                lista_titulos_blocos.append('^EDITAIS DE NOTIFICA[ÇC][AÃ]O')
                lista_titulos_blocos.append('^CHAMADA P[ÚU]BLICA')
                lista_titulos_blocos.append('^REVOGA[ÇC][ÃA]O D[AE] DISPENSA')
                lista_titulos_blocos.append('^PEDIDO DE PROPOSTA')
                   
                
                
                atualizaContadoresSecao()   
                
                #Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + Diario.get_ContadorLinhasSecao());
                
                #if (Diario.get_ContadorLinhasGeral() != list_linhas[index][12]):
                #    print('problemas')                
                #linhaFimBlocoAnterior = list_linhas[index - 1][10] - linhaFimBlocoAnterior
                #contadorLinhasGeral = contadorLinhasGeral + 1
                #contadorLinhasPagina = contadorLinhasPagina + 1
                #if (list_linhas[index - 1][2] < list_linhas[index][2]):
                    #linhaFimBlocoAnterior = 0
                    

            if pular > 0:
                pular = pular - 1
                continue
            

                                 
            #else:
                #print(bloco[index])
            #if re.compile('^INEDITORIAIS').search(each_line):
            if re.compile('^INEDIT.*').search(str(each_line).replace(' ', '')) and list_linhas[index][8] > 8:
                pular = pular + len(bloco) - index
                #Bloc.set_Fim_Bloco(str(bloco[index + w - 1]))
                break                
            
            #if re.findall('P[ÁA]GINA[ ]?\d+.*Oficial|Oficial.*P[ÁA]GINA[ ]?\d+', each_line):
            #    pagina = re.findall('P[ÁA]GINA[ ]?\d+.*Oficial|Oficial.*P[ÁA]GINA[ ]?\d+', each_line)[0]
            #    Diario.set_Pagina_Diario(re.findall('\d+',  pagina)[0])

            if len(each_line.strip()) > 0 and each_line.upper() == each_line:
                #print(each_line)
                #if re.findall('^[]?P[ÁA]GINA[ ]?\d+', each_line):
                #    Diario.set_Pagina_Diario(re.findall('\d+', each_line)[0])
                                                
                pular = montaBloco(index, each_line, pular, False, len(bloco))

                
                if re.compile(lista_Secretarias).search(each_line) and list_linhas[index][8] > 11: #re.IGNORECASE is used to ignore case    
                    nome_Secretaria = ""
                    #print(each_line + " :   Secretaria?")
                    i = 1                
                    #list = []
                    if len(str(bloco[index + i]).strip()) > 0:
                        nome_Secretaria = each_line 
                        atualizaContadoresSecretaria(i - 1, index)                
                        #Sec.set_Contador_LinhasSecretaria(contadorLinhasSecretaria_)                         
                        
                        #print(bloco[index + i])
                        #print(len(str(bloco[index + i]).strip()))                            
                        #list.append(each_line)
                        while len(str(bloco[index + i]).strip()) > 0 and str(bloco[index + i]).upper() == str(bloco[index + i]) and list_linhas[index + i][4] > deslocamento(list_linhas[index + i][9]) and list_linhas[index + i][8] > 11 and re.compile('|'.join(lista_Subsecretaria)).search(str(bloco[index + i])) is None and re.compile('|'.join(lista_Subsecretaria_Subordinados)).search(str(bloco[index + i])) is None and re.compile('|'.join(lista_titulos_blocos)).search(str(bloco[index + i])) is None:
                            nome_Secretaria = nome_Secretaria + " " + str(bloco[index + i])
                            #print(nome_Secretaria)
                            #print("passei por aqui 1")
                            #list.append(str(bloco[index + i]))
                            i = i + 1    
                            atualizaContadoresSecretaria(i - 1, index)
                            #print(nome_Secretaria) 
                            #Bloc.set_Contador_LinhasSecretaria(contadorLinhasSecretaria_)                
                    else:
                        nome_Secretaria = each_line
                        #list.append(each_line)
                    
                    #Sec.set_Secretaria(nome_Secretaria)
                    #Sec.set_Subordinada1('')
                    #Sec.set_Subordinada2('')
                    #Sec.set_Subordinada3('')
                    #pular = pular + i - 1
                    #print(nome_Secretaria)
                    
                    Sec = classe_Secretaria(nome_Secretaria)  
                    Sec.set_Contador_LinhasSecretaria(Sec.get_Contador_LinhasSecretaria() + i)
                    #Diario.set_ContadorLinhasGeral(Diario.get_ContadorLinhasGeral() + i)                           
                    #for _ in range(i):
                    #atualizaContadoresSecretaria(i - 1, index)
                        #print(nome_Secretaria)                                
                    Sec.set_Subordinada1('')
                    Sec.set_Subordinada2('')
                    Sec.set_Subordinada3('')
                    pular = pular + i - 1    
                    #print(nome_Secretaria)                
                    
                #if re.compile('^INEDITORIAIS').search(each_line):
                #    pular = pular + len(bloco) - index
                #    break
                
                #if re.compile('SUBSECRETARIA DOS CENTROS OLIMPÍCOS, PARALIMPICOS E').search(each_line):
                #    pular = pular + len(bloco) - index
                #    break                
                    #print(nome_Secretaria) #controlePrint
                    
                if re.compile('|'.join(lista_Subsecretaria)).search(each_line):                                             
                    #i = 1                        
                    #if re.compile(Sec.get_Subordinada1()).search(each_line) and len(str(bloco[index + i]).strip()) > 0 and not str(bloco[index + i]).upper() == str(bloco[index + i]):
                    #    pular = pular + i - 1
                    #else:                                                
                    #    nome_Secretaria_Subordinada1 = ""
                    nome_Secretaria_Subordinada1 = ""
                    i = 1    
                    #list = []  
                    #if len(str(bloco[index + i]).strip()) > 0:
                    if len(str(bloco[index + i]).strip()) > 0:
                        nome_Secretaria_Subordinada1 = each_line
                        atualizaContadoresSecSubordinada1(i - 1, index)
                        #Bloc.set_Contador_LinhasSecSubordinada1(contadorLinhasSecSubordinada1_)                           
                        #print(nome_Secretaria_Subordinada1)                        
                        #list.append(each_line)
                        #print(len(str(bloco[index + i]).strip()) > 0)
                        #print(re.compile('|'.join(lista_Subsecretaria)).search(str(bloco[index + i])) is None)
                        #print(re.compile('|'.join(lista_Subsecretaria_Subordinados)).search(str(bloco[index + i])) is None)
                        #print(re.compile('|'.join(lista_titulos_blocos)).search(str(bloco[index + i])) is None)                       
                        
                        while len(str(bloco[index + i]).strip()) > 0 and list_linhas[index + i][4] > deslocamento(list_linhas[index + i][9]) and checaTextoCentralizado(index, list_linhas[index + i][9]) and re.compile('|'.join(lista_Subsecretaria)).search(str(bloco[index + i])) is None and re.compile('|'.join(lista_Subsecretaria_Subordinados)).search(str(bloco[index + i])) is None and re.compile('|'.join(lista_titulos_blocos)).search(str(bloco[index + i])) is None:
                        #while len(str(bloco[index + i]).strip()) > 0 and re.compile('|'.join(lista_Subsecretaria)).search(str(bloco[index + i])) is None and re.compile('|'.join(lista_Subsecretaria_Subordinados)).search(str(bloco[index + i])) is None and re.compile('|'.join(lista_titulos_blocos)).search(str(bloco[index + i])) is None:
                        #while len(str(bloco[index + i]).strip()) > 0 and str(bloco[index + i]).upper() == str(bloco[index + i]) and list_linhas[index + i][4] > deslocamento(list_linhas[index + i][9]) and list_linhas[index + i][8] > 8: #and not re.compile('|'.join(lista_Subsecretaria)).search(str(bloco[index + i])) and not re.compile('|'.join(lista_Subsecretaria_Subordinados)).search(str(bloco[index + i])) and not re.compile('|'.join(titulos_blocos)).search(str(bloco[index + i])):                            
                            nome_Secretaria_Subordinada1 = nome_Secretaria_Subordinada1  + " " + str(bloco[index + i])
                            #list.append(str(bloco[index + i]))
                            i = i + 1    
                            atualizaContadoresSecSubordinada1(i - 1, index)  
                            #Bloc.set_Contador_LinhasSecSubordinada1(contadorLinhasSecSubordinada1_)             
                            #print(nome_Secretaria_Subordinada1) #controlePrint
                        #i = i - 1
                    else:
                        nome_Secretaria_Subordinada1 = each_line
                        #list.append(each_line)
                    #print(nome_Secretaria_Subordinada1) #controlePrint
                    #print(list)
                    #print("S1")
                                            
                    #print("SEC1 : " + nome_Secretaria_Subordinada1)
                    Sec.set_Subordinada1(nome_Secretaria_Subordinada1)
                    Sec.set_Subordinada2('')
                    Sec.set_Subordinada3('')
                    pular = pular + i - 1
                    nome_Secretaria_Subordinada2 = ""                             
                    
                    #i = i - 1
                if re.compile('|'.join(lista_Subsecretaria_Subordinados)).search(each_line) and re.compile('|'.join(lista_Subsecretaria)).search(each_line) is None:     
                    nome_Secretaria_Subordinada2 = ""
                    i = 1
                    #list = []
                    if len(str(bloco[index + i])) > 0:
                        nome_Secretaria_Subordinada2 = each_line
                        atualizaContadoresSecSubordinada2(i - 1, index)
                        #Bloc.set_Contador_LinhasSecSubordinada2(contadorLinhasSecSubordinada2_)
                        #list.append(each_line)
                        #i = i + 1
                             
                        while len(str(bloco[index + i]).strip()) > 0 and list_linhas[index + i][4] > deslocamento(list_linhas[index + i][9]) and checaTextoCentralizado(index, list_linhas[index + i][9]) and re.compile('|'.join(lista_Subsecretaria_Subordinados)).search(str(bloco[index + i])) is None and re.compile('|'.join(lista_titulos_blocos)).search(str(bloco[index + i])) is None:
                        #while len(str(bloco[index + i]).strip()) > 0 and re.compile('|'.join(lista_Subsecretaria_Subordinados)).search(str(bloco[index + i])) is None and re.compile('|'.join(lista_titulos_blocos)).search(str(bloco[index + i])) is None:
                            nome_Secretaria_Subordinada2 = nome_Secretaria_Subordinada2  + " " + str(bloco[index + i])
                            #list.append(str(bloco[index + i]))
                            i = i + 1
                            atualizaContadoresSecSubordinada2(i - 1, index)
                            #Bloc.set_Contador_LinhasSecSubordinada2(contadorLinhasSecSubordinada2_)
                            #print(nome_Secretaria_Subordinada2)
                    else:
                        nome_Secretaria_Subordinada2 = each_line
                        #list.append(each_line)
                    
                    #print("S2")
                    Sec.set_Subordinada2(nome_Secretaria_Subordinada2)
                    #print("SEC2 : " + nome_Secretaria_Subordinada2) #controlePrint
                    Sec.set_Subordinada3('')
                    #print(nome_Secretaria_Subordinada2) 
                    #pular = pular + i - 1
                    #print(pular)
                    
                        #i = i - 1
                    if (re.compile('|'.join(lista_Subsecretaria_Subordinados)).search(str(bloco[index + i])) and re.compile('|'.join(lista_Subsecretaria)).search(str(bloco[index + i])) is None):     
                        nome_Secretaria_Subordinada3 = ""
                        #list = []
                        if len(str(bloco[index + i]).strip()) > 0:
                            nome_Secretaria_Subordinada3 = str(bloco[index + i])
                            
                            #Bloc.set_Contador_LinhasSecSubordinada3(contadorLinhasSecSubordinada3_)
                            #list.append(str(bloco[index + i]))
                            i = i + 1
                            atualizaContadoresSecSubordinada3(i - 1, index)
                            #print(len(str(bloco[index + i]).strip()) > 0)
                            #print(re.compile('|'.join(lista_Subsecretaria)).search(str(bloco[index + i])) is None)
                            #print(re.compile('|'.join(lista_Subsecretaria_Subordinados)).search(str(bloco[index + i])) is None)
                            #print(re.compile('|'.join(lista_titulos_blocos)).search(str(bloco[index + i])) is None)  
                            #print(str(bloco[index + i]))                                                       
                            
                            while len(str(bloco[index + i]).strip()) > 0 and re.compile('|'.join(lista_titulos_blocos)).search(str(bloco[index + i])) is None:
                                nome_Secretaria_Subordinada3 = nome_Secretaria_Subordinada3  + " " + str(bloco[index + i])
                                #print("SEC 3 SUB: " + nome_Secretaria_Subordinada3)
                                #print(lista_titulos_blocos)
                                #list.append(str(bloco[index + i]))
                                i = i + 1
                                atualizaContadoresSecSubordinada3(i - 1, index)
                                #Bloc.set_Contador_LinhasSecSubordinada3(contadorLinhasSecSubordinada3_)
                        else:
                            nome_Secretaria_Subordinada3 = str(bloco[index + i])
                            #list.append(str(bloco[index + i]))
                        
                        #print(nome_Secretaria_Subordinada3)
                        #print("S2")
                        Sec.set_Subordinada3(nome_Secretaria_Subordinada3) 
                        #print("SEC3 : " + nome_Secretaria_Subordinada3) #controlePrint
                        #pular = pular + i - 1
                    pular = pular + i - 1
                    #print(pular)
                    #print(list_linhas[index + pular + 1][3])
        
            #texto_completo.append(string_pagina)
        

        #print(texto_completo)

        #for index, bloco in enumerate(lista_final):
            #print(bloco)    
        
        #for each in lista_final:
            #print(each)
                        
    
        wb = op.load_workbook('C:\\TesteOCR\\DODF\\Test\\OCR.xlsx')
        ws=wb['CONTRATOS']
        
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        
        #ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]|[\x00-\x1f\x7f-\x9f]|[\uffff]')
        def illegal_char_remover(data):
            """Remove ILLEGAL CHARACTER."""
            if isinstance(data, str):
                return ILLEGAL_CHARACTERS_RE.sub("", data)
            else:
                return data
    
        for idx, each_element in enumerate(lista_final):
        #for each_element in lista_excel:
            #print(each_element)
            #print(Dicionario_v1.tipo_documento_para_tipo_documento_Excel(each_element[3]))
            #ws = wb[Dicionario.tipo_documento_para_tipo_documento_Excel("CONTRATOS")]
            x = []
            for index,each in enumerate(each_element):
                #each = each.encode('unicode_escape').decode('utf-8')
                #each = each.encode('utf-8').decode('latin')
                each = illegal_char_remover(each)
                if index == 3:
                    x.append(idx + 1)
                    x.append(each)
                else:
                    x.append(each)
            ws.append(x)     
        
        #for each_element in lista_excel:
            #ws = wb[Dicionario.tipo_documento_para_tipo_documento_Excel(each_element[6])]
            #ws.append(each_element)    
        
                
        wb.save('C:\\TesteOCR\\DODF\\Test\\OCR.xlsx')  
        wb.close() 
    
        for index, each_element in enumerate(lista_final):
            
            data_blocos = {
                "NUMERO_DODF": each_element[0],
                "DATA_DODF": each_element[1],
                "COMPLEMENTO": each_element[2],
                "NUMERO_BLOCO_DODF": index + 1,                
                "SECAO": each_element[3],
                "SEC": each_element[4],
                "SUB_SEC1": each_element[5],
                "SUB_SEC2": each_element[6],
                "SUB_SEC3": each_element[7],
                "TIT_BLOCO": each_element[8],
                "SUBT_BLOCO": each_element[9],
                "DEMAIS_SUBT_BLOCO": each_element[10],
                "TEXTO_BLOCO": each_element[11],
                "INICIO_BLOCO": each_element[12],
                "FIM_BLOCO": each_element[13],
                "PAG_INICIO_BLOCO": each_element[14],
                "LINHA_INICIO_BLOCO": each_element[15],                
                "PAG_FIM_BLOCO": each_element[16],
                "LINHA_FIM_BLOCO": each_element[17],                 
                "NR_LINHAS": each_element[18],
                "CONT_SECAO": each_element[19],
                "CONT_SECREATARIA": each_element[20],     
                "CONT_SEC_SUB1": each_element[21],
                "CONT_SEC_SUB2": each_element[22],
                "CONT_SEC_SUB3": each_element[23],
                "CONT_TIT_BLOCO": each_element[24],
                "CONT_SUB_TIT_BLOCO": each_element[25],      
                "CONT_DEMAIS_SUBT_BLOCO": each_element[26],
                "CONT_TOTAL_BLOCO": each_element[27],                  
                                
            }            
            #db_connection.insert_dodf_bloco(connection, data_blocos)

    #connection.close()
    #buscaSimilaridades
#run_program([['C:\\TesteOCR\\DODF\\DODF downloads\\2018\\07\\DODF 137 20-07-2018 INTEGRA.pdf', 'DODF 137 20-07-2018 INTEGRA.pdf']])
#run_program([['C:\\TesteOCR\\DODF\\DODF downloads\\2018\\07\\DODF 139 24-07-2018 INTEGRA.pdf', 'DODF 139 24-07-2018 INTEGRA.pdf']])
#run_program([['C:\\TesteOCR\\DODF\\DODF downloads\\2018\\07\\DODF 141 26-07-2018 INTEGRA.pdf', 'DODF 141 26-07-2018 INTEGRA.pdf']])
#DODF 141 26-07-2018 INTEGRA.pdf
#DODF 124 03-07-2018 INTEGRA.pdf

#run_program([['C:\\TesteOCR\\DODF\\DODF downloads\\2019\\01\\DODF 005 08-01-2019 INTEGRA.pdf', 'DODF 005 08-01-2019 INTEGRA.pdf']])

#run_program([['C:\\TesteOCR\\DODF\\DODF downloads\\2018\\11\\DODF 212 07-11-2018 INTEGRA.pdf', 'DODF 212 07-11-2018 INTEGRA.pdf']])
#run_program(None,"C:/TesteOCR/DODF/DODF downloads/2018/11_v3")
#run_program(None,"C:/TesteOCR/DODF/DODF downloads/2018/11")
run_program(None,"C:/TesteOCR/DODF/DODF downloads/2019/03")
#run_program(None,"C:/TesteOCR/DODF/DODF downloads/2018/11")